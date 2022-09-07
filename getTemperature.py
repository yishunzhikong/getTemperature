import requests
import re
import os
import time
import datetime
import random
import openpyxl
from openpyxl import load_workbook

#test

#必改数据
#查询城市
City = "beijing"
#查询开始&结束时间
timeBegin = "20190815"
timeEnd = "20220815"

# URL常量
website = "https://www.tianqi.com"
Temperature = '<span>(-?[0-9]*?)~(-?[0-9]*?)°</span>'
UserAgent = {
                'User-Agent':
                    'Mozilla/5.0 (Windows NT 10.0; Win64; x64) \
                    AppleWebKit/537.36 (KHTML, like Gecko) \
                    Chrome/99.0.4844.84 Safari/537.36'
             }

#获取某个城市某天温度
def getTemp( city, date):
    s = requests.session()
    rs = s.get(website + "/tianqi/" + city + "/" + date + ".html", headers=UserAgent)
    temp = re.search(Temperature, rs.text, re.S)
    return date,int(temp.group(1)),int(temp.group(2))

# 主函数
if __name__ == '__main__':

    i = 2

    #限定时间范围
    begin = datetime.date( \
        int(timeBegin[0:4]), \
        int(timeBegin[4:6]), \
        int(timeBegin[6:8]))
    end = datetime.date( \
        int(timeEnd[0:4]), \
        int(timeEnd[4:6]), \
        int(timeEnd[6:8]))
    delta = datetime.timedelta(days=1)

    #打开excel表格
    tempWb = openpyxl.Workbook()
    tempWs = tempWb.active
    
    #写入表头
    tempWs.cell(1, 1, value="date")
    tempWs.cell(1, 2, value="min")
    tempWs.cell(1, 3, value="max")

    while begin <= end:
        nowdate = begin.strftime("%Y%m%d")
        tp = getTemp( City, nowdate)
        tempWs.cell(i, 1, value=tp[0])
        tempWs.cell(i, 2, value=tp[1])
        tempWs.cell(i, 3, value=tp[2])
        begin += delta
        time.sleep(random.randint(5,20)/10)
        i += 1
        #debug
        print(tp)

    #保存表格文件
    tempWb.save(City + "_" + timeBegin + "_" +timeEnd + ".xlsx")